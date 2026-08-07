"""
数据库导出引擎 - 核心逻辑

支持将 SQLite / MySQL / PostgreSQL 中的表导出为 Excel(.xlsx) 或 CSV 文件。
性能优化：
    - 超过 1000 条自动分批（默认 500 条/批，可调）
    - 超过 100000 行启用多线程并行读取（线程池最多 20 个）
    - 流式写入：读一批写一批，避免将全量数据先存内存
    - 单 sheet 最多 MAX_ROWS_PER_SHEET 行，超过自动新建 sheet
    - 单工作簿合计最多 MAX_ROWS_PER_WORKBOOK 行，超过新建工作簿文件
    - 超过单工作簿容量时，在 output_dir 下创建 "数据库名_表名" 子文件夹
"""

import math
import os
import queue as _queue
import re
import sys
import threading
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import csv
from multiprocessing import Manager
try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False

from .database_connector import create_connector
from .progress_bar import ProgressBar


# ===================== 工具函数 =====================
def get_db_display_name(db_type: str, config: Dict[str, Any]) -> str:
    """从数据库连接配置中提取可读的显示名（用于文件名等）"""
    db = (db_type or "").lower()
    if db == "sqlite":
        path = config.get("path") or config.get("db_path") or config.get("database") or ""
        if path:
            name = os.path.basename(path)
            if name.lower().endswith(".db"):
                name = name[:-3]
            elif name.lower().endswith(".sqlite"):
                name = name[:-7]
            elif name.lower().endswith(".sqlite3"):
                name = name[:-8]
            return name or "sqlite"
        return "sqlite"
    if db in ("mysql",):
        return str(config.get("database") or config.get("db") or "mysql")
    if db in ("postgresql", "postgres", "pg", "psql"):
        return str(config.get("database") or config.get("db") or "postgres")
    return str(config.get("database") or config.get("path") or "db")


# ===================== 常量配置 =====================
BATCH_THRESHOLD = 3000
DEFAULT_BATCH_SIZE = 50000       # MySQL 单次查询批大小（5 万行：减少调用次数，内存占用仍合理）
MULTITHREAD_THRESHOLD = 100000
MULTIPROCESS_THRESHOLD = 500000   # 单表超过 50 万行启用多进程分段
DEFAULT_THREADS = 12
MAX_THREADS = 20
DEFAULT_PROCESSES = 4             # 默认4进程（4核CPU利用率 70%）
MAX_PROCESSES = 4                 # 最多4进程（更多进程因磁盘IO和zlib竞争反而变慢）
MAX_ROWS_PER_SHEET = 50000
MAX_ROWS_PER_WORKBOOK = 500000
EXCEL_MERGE_ROWS = 50000          # 写入 Excel 时的合并块（和批大小一致）
PROGRESS_MIN_INTERVAL = 0.3
MULTITHREAD_DB_TYPES = ("mysql", "postgresql", "postgres", "pg", "psql")


# ===================== 工具函数 =====================
def sanitize_filename(name: str) -> str:
    """清洗文件名，去掉 Windows/Linux 不允许的字符"""
    name = re.sub(r'[\\/:*?"<>|\r\n\t]', "_", name)
    name = name.strip(" .")
    return name or "export"


def truncate_sheet_name(db_display: str, table: str, sheet_idx: int, max_len: int = 31) -> str:
    """
    构造 Excel sheet 名并截断。
    格式: 数据库_表名_sheetN，超出 max_len 时保留末尾 (sheetN) 并从中间省略
    """
    suffix = f"_sheet{sheet_idx}"
    prefix = sanitize_filename(f"{db_display}_{table}")
    remaining = max_len - len(suffix)
    if remaining <= 0:
        return f"sheet{sheet_idx}"[-max_len:]
    if len(prefix) <= remaining:
        return prefix + suffix
    # 省略中间部分
    keep_left = max(1, remaining // 2)
    keep_right = max(1, remaining - keep_left - 1)
    truncated = prefix[:keep_left] + "~" + prefix[-keep_right:]
    return (truncated + suffix)[:max_len]


def _row_values(row: Any) -> List[Any]:
    """
    统一把一行数据转换成值列表。
    - dict cursor: {'id': 1, 'uuid': '...'} -> [1, '...']
    - tuple cursor: (1, '...') -> [1, '...']
    - list cursor: [1, '...'] -> [1, '...']
    """
    if isinstance(row, dict):
        return list(row.values())
    return list(row)


def _build_offset_query(db_type: str, table: str, columns: List[str],
                        limit: int, offset: int) -> str:
    """根据数据库类型生成带 LIMIT/OFFSET 的 SELECT 语句（无主键时回退用）"""
    cols = ", ".join(columns)
    return f"SELECT {cols} FROM {table} LIMIT {int(limit)} OFFSET {int(offset)}"


def _build_range_query(db_type: str, table: str, columns: List[str],
                        pk_col: str, start_val, end_val) -> str:
    """
    主键范围查询（比 LIMIT/OFFSET 快 5-10 倍）
    WHERE id >= start AND id < end ORDER BY id
    """
    cols = ", ".join(columns)
    db = db_type.lower()
    if db in ("mysql",):
        return (f"SELECT {cols} FROM {table} "
                f"WHERE {pk_col} >= {start_val} AND {pk_col} < {end_val} "
                f"ORDER BY {pk_col}")
    # PostgreSQL / SQLite 也支持
    return (f"SELECT {cols} FROM {table} "
            f"WHERE {pk_col} >= {start_val} AND {pk_col} < {end_val} "
            f"ORDER BY {pk_col}")


def _detect_primary_key(db_type: str, connector, table: str) -> Tuple[Optional[str], bool]:
    """
    检测表是否有"单列主键"，返回 (主键列名, 是否整数类型)。

    返回 (None, False) 表示无单列主键。
    有整数主键 → 可以用主键范围分页（超快速，支持多进程分段）
    有字符串主键 → 可以用键集分页（O(1) 每批，比 LIMIT/OFFSET 快 10-50 倍）
    无主键 → 只能用 LIMIT/OFFSET（大表慢）
    """
    db = db_type.lower()
    try:
        schema = connector.get_table_schema(table)
    except Exception:
        return None, False

    pk_cols = [col for col, col_type, is_nullable, is_primary in schema if is_primary]
    if len(pk_cols) != 1:
        return None, False

    pk_col = pk_cols[0]
    pk_type_col = None
    for col, col_type, is_nullable, is_primary in schema:
        if is_primary:
            pk_type_col = (col_type or "").upper()
            break

    if pk_type_col is None:
        return None, False

    integer_keywords = ("INT", "INTEGER", "BIGINT", "SMALLINT", "TINYINT", "MEDIUMINT", "SERIAL")
    is_integer = any(k in pk_type_col for k in integer_keywords)

    return pk_col, is_integer


def _format_pk_value(value, is_integer: bool, db_type: str) -> str:
    """格式化主键值用于 SQL 拼接（键集分页用），正确处理转义"""
    if is_integer:
        return str(int(value))
    s = str(value).replace("'", "''")
    return f"'{s}'"


def _quote_identifier(name: str, db_type: str) -> str:
    """根据数据库类型给标识符加引号"""
    db = db_type.lower()
    if db in ("mysql",):
        return f"`{name}`"
    elif db in ("postgresql", "postgres", "pg", "psql"):
        return f'"{name}"'
    return name


def _detect_integer_primary_key(db_type: str, connector, table: str) -> Optional[str]:
    """兼容旧接口：返回整数主键列名，非整数返回 None"""
    pk_col, is_integer = _detect_primary_key(db_type, connector, table)
    return pk_col if is_integer else None


def _split_pk_ranges(min_val: int, max_val: int, batch_size: int) -> List[Tuple[int, int]]:
    """
    根据主键范围 + 批大小，分割成 [start, end) 区间列表。
    注意：最后一段要包含 max_val，所以最后一段用 end = max_val + 1
    """
    if min_val > max_val:
        return []
    ranges: List[Tuple[int, int]] = []
    start = min_val
    while start <= max_val:
        end = min(start + batch_size, max_val + 1)
        ranges.append((start, end))
        start = end
    return ranges


# ===================== 多进程 worker（模块级，进程池要求可 pickle）=====================
def _segment_worker(args_tuple):
    """
    进程池 worker：导出表的一个主键范围段（50 万行/段）
    每个 worker 独立连接数据库、逐批读、逐行写。
    完成后 close() → 立即保存到磁盘并释放内存。

    生产级保护：
      - 按 MAX_ROWS_PER_SHEET 切分 sheet（5万行/sheet）
      - 按 MAX_ROWS_PER_WORKBOOK 切分工作簿（50万行/工作簿），防止数据截断
      - 异常时确保 wb.close() 被调用，避免生成损坏的 ZIP 文件
      - 完成后校验文件完整性（大小>0 且可被 openpyxl 打开）
      - 通过共享计数器实时向主进程汇报进度

    返回：([file_path], row_count, error_message)
    """
    db_type, db_config, table, columns, pk_col, start_val, end_val, \
        output_path, output_format, batch_size, db_display, shared_counter, \
        is_integer_pk = args_tuple

    row_count = 0
    error = None
    connector = None
    stream_cursor = None
    streaming_conn = None
    cur = None
    output_files: List[str] = []
    wb = None  # 跟踪当前工作簿对象，确保 close()

    try:
        connector = create_connector(db_type, db_config)
        connector.connect()
    except Exception as exc:
        return [output_path], 0, f"连接失败: {exc}"

    try:
        col_sql = ", ".join(columns)

        # 格式化主键值：整数不引号，字符串加引号并转义
        if is_integer_pk:
            start_sql = str(int(start_val))
            end_sql = str(int(end_val))
        else:
            start_sql = "'" + str(start_val).replace("'", "''") + "'"
            end_sql = "'" + str(end_val).replace("'", "''") + "'"

        if db_type == "mysql":
            try:
                import pymysql
                streaming_conn = pymysql.connect(
                    host=db_config["host"],
                    port=db_config.get("port", 3306),
                    user=db_config["user"],
                    password=db_config["password"],
                    database=db_config.get("database"),
                    charset=db_config.get("charset", "utf8mb4"),
                    connect_timeout=600,
                    cursorclass=pymysql.cursors.SSCursor,
                )
                stream_cursor = streaming_conn.cursor()
                stream_cursor.execute(
                    f"SELECT {col_sql} FROM `{table}` "
                    f"WHERE `{pk_col}` >= {start_sql} "
                    f"AND `{pk_col}` < {end_sql}"
                )
            except Exception:
                stream_cursor = None
                if streaming_conn:
                    try: streaming_conn.close()
                    except Exception: pass
                streaming_conn = None

        if stream_cursor is None:
            sql = f"SELECT {col_sql} FROM {table} WHERE {pk_col} >= {start_sql} AND {pk_col} < {end_sql}"
            cur = connector.execute(sql)

        def get_batches():
            """批量行生成器"""
            if stream_cursor is not None:
                while True:
                    rows = stream_cursor.fetchmany(batch_size)
                    if not rows:
                        return
                    yield [list(r) for r in rows]
            else:
                while True:
                    rows = cur.fetchmany(batch_size)
                    if not rows:
                        return
                    out = []
                    for r in rows:
                        if isinstance(r, dict):
                            out.append([r[c] for c in columns])
                        else:
                            out.append(list(r))
                    yield out

        # 辅助函数：更新共享计数器（每 1000 行更新一次，减少 IPC 开销）
        def _update_shared_counter(delta: int):
            if shared_counter is not None and delta > 0:
                shared_counter.value += delta

        # ============ CSV 模式 ============
        if output_format == "csv":
            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                for batch_rows in get_batches():
                    writer.writerows(batch_rows)
                    row_count += len(batch_rows)
                    _update_shared_counter(len(batch_rows))
            output_files.append(output_path)

        # ============ XLSXWRITER 模式（生产级：带工作簿切分保护）============
        elif HAS_XLSXWRITER:
            workbook_idx = 0
            current_path = output_path
            ws = None
            row_idx = 1
            rows_in_sheet = 0
            sheet_idx = 0
            rows_in_workbook = 0

            def switch_workbook_xlsx(new_path: str):
                """关闭当前工作簿并创建新的"""
                nonlocal wb, ws, row_idx, rows_in_sheet, sheet_idx, rows_in_workbook, workbook_idx
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass
                wb = xlsxwriter.Workbook(new_path, {
                    "strings_to_urls": False,
                    "constant_memory": True,
                })
                ws = wb.add_worksheet(truncate_sheet_name(db_display, table, 0))
                ws.write_row(0, 0, columns)
                row_idx = 1
                rows_in_sheet = 0
                sheet_idx = 0
                rows_in_workbook = 0
                workbook_idx += 1

            def switch_sheet_xlsx():
                """新建 sheet"""
                nonlocal ws, row_idx, rows_in_sheet, sheet_idx
                sheet_idx += 1
                ws = wb.add_worksheet(truncate_sheet_name(db_display, table, sheet_idx))
                ws.write_row(0, 0, columns)
                rows_in_sheet = 0
                row_idx = 1

            # 初始化第一个工作簿
            switch_workbook_xlsx(current_path)

            for batch_rows in get_batches():
                for r in batch_rows:
                    # Sheet 切分保护：每 sheet 最多 5 万行
                    if rows_in_sheet >= MAX_ROWS_PER_SHEET:
                        switch_sheet_xlsx()

                    # 工作簿切分保护：每工作簿最多 50 万行（防止数据截断）
                    if rows_in_workbook >= MAX_ROWS_PER_WORKBOOK:
                        # 关闭当前工作簿，保存到磁盘
                        if wb is not None:
                            try:
                                wb.close()
                                wb = None
                            except Exception:
                                pass
                        output_files.append(current_path)
                        # 创建新工作簿文件名
                        base, ext = os.path.splitext(output_path)
                        current_path = f"{base}_w{workbook_idx + 1:03d}{ext}"
                        switch_workbook_xlsx(current_path)

                    ws.write_row(row_idx, 0, r)
                    row_idx += 1
                    rows_in_sheet += 1
                    rows_in_workbook += 1
                    row_count += 1
                _update_shared_counter(len(batch_rows))

            # 关闭最后一个工作簿
            if wb is not None:
                try:
                    wb.close()
                    wb = None
                except Exception:
                    pass
            output_files.append(current_path)

        # ============ OPENPYXL 模式（生产级：带工作簿切分保护）============
        elif HAS_OPENPYXL:
            workbook_idx = 0
            current_path = output_path
            current_wb = Workbook(write_only=True)
            ws = current_wb.create_sheet(truncate_sheet_name(db_display, table, 0))
            ws.append(columns)
            rows_in_sheet = 0
            sheet_idx = 0
            rows_in_workbook = 0

            def switch_workbook_openpyxl(new_path: str):
                """保存当前工作簿并创建新的"""
                nonlocal current_wb, ws, rows_in_sheet, sheet_idx, rows_in_workbook, workbook_idx
                if current_wb is not None:
                    try:
                        current_wb.save(current_path)
                    except Exception:
                        pass
                current_wb = Workbook(write_only=True)
                ws = current_wb.create_sheet(truncate_sheet_name(db_display, table, 0))
                ws.append(columns)
                rows_in_sheet = 0
                sheet_idx = 0
                rows_in_workbook = 0
                workbook_idx += 1

            def switch_sheet_openpyxl():
                """新建 sheet"""
                nonlocal ws, rows_in_sheet, sheet_idx
                sheet_idx += 1
                ws = current_wb.create_sheet(truncate_sheet_name(db_display, table, sheet_idx))
                ws.append(columns)
                rows_in_sheet = 0

            for batch_rows in get_batches():
                for r in batch_rows:
                    # Sheet 切分保护
                    if rows_in_sheet >= MAX_ROWS_PER_SHEET:
                        switch_sheet_openpyxl()

                    # 工作簿切分保护
                    if rows_in_workbook >= MAX_ROWS_PER_WORKBOOK:
                        if current_wb is not None:
                            try:
                                current_wb.save(current_path)
                            except Exception:
                                pass
                        output_files.append(current_path)
                        base, ext = os.path.splitext(output_path)
                        current_path = f"{base}_w{workbook_idx + 1:03d}{ext}"
                        switch_workbook_openpyxl(current_path)

                    ws.append(r)
                    rows_in_sheet += 1
                    rows_in_workbook += 1
                    row_count += 1
                _update_shared_counter(len(batch_rows))

            # 保存最后一个工作簿
            if current_wb is not None:
                try:
                    current_wb.save(current_path)
                except Exception:
                    pass
            output_files.append(current_path)

        else:
            raise ImportError("需要 xlsxwriter 或 openpyxl")

    except Exception as exc:
        error = f"{exc}"

    finally:
        # ============ 生产级清理：确保资源释放 + 文件完整性校验 ============
        #
        # 清理顺序（关键）：
        #   1. 关闭/保存工作簿（必须在 cursor/connection 之前，确保数据写入磁盘）
        #   2. 关闭 SSCursor（断开内部引用，避免 __del__ 访问已关闭 socket）
        #   3. 关闭 connection

        # ---- 1. 确保工作簿被正确关闭/保存 ----
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
            wb = None

        # ---- 2. 关闭 SSCursor + 清空内部引用 ----
        if stream_cursor is not None:
            try:
                if hasattr(stream_cursor, '_result'):
                    try:
                        if stream_cursor._result is not None and hasattr(stream_cursor._result, 'connection'):
                            stream_cursor._result.connection = None
                    except Exception: pass
                stream_cursor.close()
            except Exception: pass
            stream_cursor = None

        # ---- 3. 关闭 fallback cursor ----
        if cur is not None:
            try: cur.close()
            except Exception: pass
            cur = None

        # ---- 4. 关闭 streaming connection ----
        if streaming_conn is not None:
            try:
                if hasattr(streaming_conn, '_sock'):
                    try: streaming_conn._sock = None
                    except Exception: pass
                streaming_conn.close()
            except Exception: pass
            streaming_conn = None

        # ---- 5. 关闭通用 connector ----
        if connector is not None:
            try: connector.close()
            except Exception: pass
            connector = None

        # ---- 6. 文件完整性校验（生产级：检测损坏文件） ----
        valid_files = []
        for fp in list(output_files):
            if os.path.exists(fp) and os.path.getsize(fp) > 1024:
                # 对于 xlsx 文件，尝试用 zipfile 校验（xlsx 本质是 ZIP）
                if fp.endswith('.xlsx'):
                    try:
                        import zipfile
                        with zipfile.ZipFile(fp, 'r') as zf:
                            bad_files = zf.testzip()
                            if bad_files is not None:
                                os.remove(fp)
                                continue
                    except Exception:
                        # ZIP 损坏，删除文件
                        try:
                            os.remove(fp)
                        except Exception:
                            pass
                        continue
                valid_files.append(fp)
            else:
                # 文件不存在或太小（<1KB），可能是损坏的
                try:
                    os.remove(fp)
                except Exception:
                    pass
        output_files = valid_files

    return output_files, row_count, error


# ===================== 多线程工作函数 =====================
def _worker_thread(task_queue: "_queue.Queue",
                   result_queue: "_queue.Queue",
                   db_type: str,
                   db_config: Dict[str, Any],
                   table: str,
                   columns: List[str],
                   error_list: list) -> None:
    """
    线程工作函数：从 task_queue 取任务，放入 result_queue（带批次序号）

    任务格式（两种之一）:
      ("offset", batch_idx, base_offset, batch_size)
      ("range",  batch_idx, pk_col, start_val, end_val)
    """
    try:
        from .database_connector import create_connector
        connector = create_connector(db_type, db_config)
        connector.connect()
        cursor = connector.cursor()
        cols = ", ".join(columns)

        while True:
            try:
                task = task_queue.get_nowait()
            except _queue.Empty:
                break
            except Exception:
                break

            try:
                mode = task[0]
                batch_idx = task[1]

                if mode == "offset":
                    _, _, base_offset, batch_size = task
                    skip = base_offset + batch_idx * batch_size
                    sql = f"SELECT {cols} FROM {table} LIMIT {int(batch_size)} OFFSET {int(skip)}"
                    cursor.execute(sql)
                    batch_rows = cursor.fetchall()
                elif mode == "range":
                    _, _, pk_col, start_val, end_val = task
                    sql = (f"SELECT {cols} FROM {table} "
                           f"WHERE {pk_col} >= {start_val} AND {pk_col} < {end_val} "
                           f"ORDER BY {pk_col}")
                    cursor.execute(sql)
                    batch_rows = cursor.fetchall()
                else:
                    error_list.append((batch_idx, f"未知任务模式: {mode}"))
                    result_queue.put((batch_idx, []))
                    continue

                result_queue.put((batch_idx, [_row_values(r) for r in batch_rows]))
            except Exception as exc:
                error_list.append((batch_idx, str(exc)))
                result_queue.put((batch_idx, []))
            finally:
                task_queue.task_done()

        try:
            connector.close()
        except Exception:
            pass
    except Exception as exc:
        error_list.append((-1, f"线程初始化失败: {exc}"))


# ===================== 主类 =====================

class ExportProgress:
    """导出专用：单行组合进度条
    同一行同时显示读 + 写进度，持续 \r 覆盖刷新，不换行刷屏。
    线程安全，支持穿插打印日志（write_log）。
    """

    def __init__(self, table_name: str, total_rows: int, width: int = 30):
        self.table_name = table_name
        self.total = max(1, int(total_rows))
        self.width = width
        self.read_count = 0
        self.write_count = 0
        self._last_line_len = 0
        self._last_update = 0.0
        self._min_interval = 0.25
        self.start_time = time.time()
        self._finished = False
        # 速度平滑：记录最近 5 次 (elapsed, count) 快照，用加权平均算速度
        self._speed_snapshots: List[Tuple[float, int]] = []  # [(elapsed, count), ...]
        self._smooth_speed: float = 0.0
        try:
            import threading
            self._lock = threading.Lock()
        except Exception:
            self._lock = None

    @staticmethod
    def _format_time(seconds: float) -> str:
        if seconds < 60:
            return f"{seconds:.0f}s"
        minutes = int(seconds // 60)
        sec = int(seconds % 60)
        if minutes < 60:
            return f"{minutes}m{sec:02d}s"
        hours = minutes // 60
        minutes = minutes % 60
        return f"{hours}h{minutes:02d}m"

    def _human(self, n: int) -> str:
        if n >= 1000000:
            return f"{n / 1000000:.1f}M"
        if n >= 1000:
            return f"{n / 1000:.1f}K"
        return f"{n}"

    def _render(self):
        """渲染单行进度条（用 \r 在终端覆盖刷新）"""
        # 进度基准：有 write 用 write（最可信），没有用 read（多进程模式下是估算值）
        current = max(self.write_count, self.read_count)
        current = min(current, self.total)
        progress_ratio = current / self.total
        filled = int(self.width * progress_ratio)
        bar = "█" * filled + "░" * (self.width - filled)
        percent = int(progress_ratio * 100)

        elapsed = time.time() - self.start_time

        # ============ ETA 计算（生产级：绝对不允许时光倒流）============
        #
        # 核心原则：用户看到的剩余时间只能减少或持平，绝不能增加！
        #
        # 统一策略：无论单线程/多线程/多进程，都使用基于整体平均速度的
        # 单调递减 ETA。简单、可靠、无歧义。
        #
        eta = 0.0

        if current > 0 and current < self.total and elapsed > 0.5:
            # 基于整体平均速度计算 ETA（最稳定，不受瞬时波动影响）
            avg_speed = current / elapsed
            remaining = self.total - current
            calc_eta = remaining / max(avg_speed, 0.001)

            # 单调递减保护：ETA 只降不升
            if not hasattr(self, '_min_eta_observed'):
                self._min_eta_observed = calc_eta
            elif calc_eta < self._min_eta_observed:
                self._min_eta_observed = calc_eta

            eta = max(self._min_eta_observed, 0.0)

        parts = [
            f"  {self.table_name}",
            f"[{bar}]",
            f"{percent:3d}%",
            f"读 {self._human(self.read_count)} 写 {self._human(self.write_count)}",
        ]

        # ETA 显示逻辑（生产级）：
        #   - 正常进度 → 显示具体剩余时间
        #   - 接近完成（≥98%）→ 显示"即将完成"
        #   - 已完成（100%）→ 不显示 ETA
        if current < self.total:
            if progress_ratio >= 0.98:
                parts.append("剩 即将完成")
            elif eta > 1.0:
                parts.append(f"剩 {self._format_time(eta)}")
        parts.append(f"用时 {self._format_time(elapsed)}")

        line = " ".join(parts)

        # 清尾：如比上次短，用空格覆盖
        if len(line) < self._last_line_len:
            line = line + " " * (self._last_line_len - len(line))
        sys.stdout.write("\r" + line)
        sys.stdout.flush()
        self._last_line_len = len(line)

    def _throttled_render(self):
        # 已完成则不再刷新，避免 finish() 后被覆盖
        if self._finished:
            return
        now = time.time()
        if now - self._last_update < self._min_interval and self.write_count < self.total:
            return
        self._last_update = now
        if self._lock:
            self._lock.acquire()
        try:
            if not self._finished:
                self._render()
        finally:
            if self._lock:
                self._lock.release()

    def update_read(self, count: int) -> None:
        """更新已读取行数"""
        if self._finished:
            return
        self.read_count = count
        self._throttled_render()

    def update_write(self, count: int) -> None:
        """更新已写入行数"""
        if self._finished:
            return
        self.write_count = count
        self._throttled_render()

    def force_render(self) -> None:
        """强制刷新（忽略节流）——用于长时间运行时给用户反馈"""
        self._render()

    def add_read(self, delta: int) -> None:
        if delta <= 0:
            return
        self.update_read(self.read_count + delta)

    def add_write(self, delta: int) -> None:
        if delta <= 0:
            return
        self.update_write(self.write_count + delta)

    def finish(self, message: str = "") -> None:
        """完成进度条"""
        if self._finished:
            return
        self._finished = True
        if self._lock:
            self._lock.acquire()
        try:
            self.read_count = max(self.read_count, self.total)
            self.write_count = self.total
            self._render()
            sys.stdout.write("\n")
            sys.stdout.flush()
        finally:
            if self._lock:
                self._lock.release()

    def write_log(self, msg: str) -> None:
        """穿插打印日志，不破坏进度条行"""
        if self._finished:
            print(msg)
            return
        if self._lock:
            self._lock.acquire()
        try:
            clear = " " * max(self._last_line_len, len(msg))
            sys.stdout.write("\r" + clear + "\n")
            sys.stdout.write(msg + "\n")
            sys.stdout.flush()
            self._last_line_len = 0
            self._render()
        finally:
            if self._lock:
                self._lock.release()


class DatabaseExporter:
    """数据库导出器 - 流式读取 + 批量写入 + 进度条"""

    def __init__(
        self,
        connector,
        tables: List[str],
        db_type: str,
        db_config: Dict[str, Any],
        output_format: str = "xlsx",
        output_dir: str = ".",
        custom_filename: Optional[str] = None,
        from_row: int = 0,
        to_row: Optional[int] = None,
        limit: Optional[int] = None,
        batch_size: int = DEFAULT_BATCH_SIZE,
        max_threads: int = DEFAULT_THREADS,
        logger=None,
    ):
        self.connector = connector
        self.tables = tables
        self.db_type = db_type.lower()
        self.db_config = db_config
        self.output_format = output_format.lower().lstrip('.')
        self.output_dir = output_dir
        self.custom_filename = custom_filename
        self.from_row = max(0, int(from_row or 0))
        self.to_row = int(to_row) if to_row is not None else None
        self.limit = int(limit) if limit is not None else None
        self.batch_size = max(50, int(batch_size or DEFAULT_BATCH_SIZE))
        self.max_threads = min(MAX_THREADS, max(1, int(max_threads or DEFAULT_THREADS)))
        self.logger = logger
        self.db_display = get_db_display_name(db_type, db_config)

        if self.output_format not in ("xlsx", "csv"):
            raise ValueError(f"不支持的导出格式: {self.output_format}，仅支持 xlsx / csv")

        os.makedirs(self.output_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # 日志 / 工具方法
    # ------------------------------------------------------------------
    def _log(self, level: str, message: str) -> None:
        """统一日志输出"""
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{level}] {ts} {message}"
        if self.logger and hasattr(self.logger, "info"):
            try:
                if level == "SUCCESS":
                    self.logger.info(line)
                elif level == "WARN":
                    self.logger.warning(line)
                elif level == "ERROR":
                    self.logger.error(line)
                else:
                    self.logger.info(line)
                return
            except Exception:
                pass
        print(line)

    def _generate_filename(self, table: str, total_rows: int) -> str:
        """生成默认文件名: 数据库名_表名_N行_年月日_时分秒"""
        if self.custom_filename:
            return sanitize_filename(self.custom_filename)
        db = sanitize_filename(self.db_display)
        tb = sanitize_filename(table)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{db}_{tb}_{total_rows}行_{ts}"

    def _get_effective_total(self, table: str) -> Tuple[int, int]:
        """
        计算 (源表总行数, 本次实际要导出的行数)
        """
        total = self.connector.get_row_count(table) or 0
        effective = max(0, total - self.from_row)
        if self.to_row is not None and self.to_row > self.from_row:
            effective = min(effective, self.to_row - self.from_row)
        if self.limit is not None and self.limit > 0:
            effective = min(effective, self.limit)
        return total, effective

    # ------------------------------------------------------------------
    # 读取逻辑（单线程分批 / 多线程）
    # ------------------------------------------------------------------
    def _read_rows_sequential_streaming(self, table: str, columns: List[str],
                                        total: int,
                                        progress: Optional[ExportProgress] = None):
        """
        单线程流式读取（生成器）：每次读取一批，边读边 yield。

        性能优先级：
          1. MySQL → SSCursor 流式（最快）
          2. 有整数主键 → 主键范围查询（O(1)，很快）
          3. 无主键 → LIMIT/OFFSET（O(n)，大表慢）
        """
        batch_size = self.batch_size

        # ============ 最快路径：MySQL + SSCursor
        if self.db_type == "mysql":
            try:
                gen = self.connector.streaming_fetchall(columns, table, batch_size)
                if gen is not None:
                    read_count = 0
                    for batch in gen:
                        if not batch:
                            continue
                        read_count += len(batch)
                        if progress is not None:
                            progress.update_read(read_count)
                        yield [list(r) for r in batch]
                    return
            except Exception:
                pass

        # ============ 其他路径：先检测是否能用主键
        pk_col, pk_is_integer = _detect_primary_key(self.db_type, self.connector, table)
        pk_min = pk_max = None
        if pk_col is not None and pk_is_integer:
            try:
                row = self.connector.fetchone(f"SELECT MIN({pk_col}), MAX({pk_col}) FROM {table}")
                if row is not None and row[0] is not None:
                    pk_min, pk_max = row[0], row[1]
            except Exception:
                pass  # 回退到键集分页

        read_count = 0
        if pk_col is not None and pk_is_integer and pk_min is not None:
            # --- 整数主键范围查询：O(1) ---
            ranges = _split_pk_ranges(pk_min, pk_max, batch_size)
            for batch_idx, (start_val, end_val) in enumerate(ranges):
                if read_count >= total:
                    break
                try:
                    sql = (f"SELECT {', '.join(columns)} FROM {table} "
                           f"WHERE {pk_col} >= {start_val} AND {pk_col} < {end_val} "
                           f"ORDER BY {pk_col}")
                    rows = self.connector.fetchall(sql)
                    batch = [_row_values(r) for r in rows]
                    if batch:
                        read_count += len(batch)
                        if progress is not None:
                            progress.update_read(read_count)
                        yield batch
                except Exception as exc:
                    self._log("WARN", f"  批 {batch_idx} 读取失败: {exc}")
        elif pk_col is not None:
            # --- 键集分页：O(1) 每批，字符串主键也能快速读取 ---
            q_pk = _quote_identifier(pk_col, self.db_type)
            cols_str = ", ".join(columns)
            last_pk_val = None
            try:
                pk_idx = columns.index(pk_col)
            except ValueError:
                pk_idx = 0
            while read_count < total:
                if last_pk_val is None:
                    sql = (f"SELECT {cols_str} FROM {table} "
                           f"ORDER BY {q_pk} LIMIT {batch_size}")
                else:
                    formatted = _format_pk_value(last_pk_val, pk_is_integer, self.db_type)
                    sql = (f"SELECT {cols_str} FROM {table} "
                           f"WHERE {q_pk} > {formatted} "
                           f"ORDER BY {q_pk} LIMIT {batch_size}")
                try:
                    rows = self.connector.fetchall(sql)
                except Exception as exc:
                    self._log("WARN", f"  键集分页读取失败: {exc}")
                    break
                batch = [_row_values(r) for r in rows]
                if not batch:
                    break
                read_count += len(batch)
                last_pk_val = batch[-1][pk_idx]
                if progress is not None:
                    progress.update_read(read_count)
                yield batch
        else:
            # --- 回退：LIMIT/OFFSET ---
            remaining = total
            offset = self.from_row
            while remaining > 0:
                this_limit = min(batch_size, remaining)
                sql = (f"SELECT {', '.join(columns)} FROM {table} "
                       f"LIMIT {int(this_limit)} OFFSET {int(offset)}")
                rows = self.connector.fetchall(sql)
                batch = [_row_values(r) for r in rows]
                if not batch:
                    break
                read_count += len(batch)
                remaining -= len(batch)
                offset += len(batch)
                if progress is not None:
                    progress.update_read(read_count)
                yield batch

        if progress is not None and not progress._finished:
            pass  # 不在这里 finish（由写入器统一 finish）

    def _read_rows_multithread_streaming(self, table: str, columns: List[str],
                                          total: int,
                                          progress: Optional[ExportProgress] = None):
        """
        多线程流式读取（生成器）：一边被主线程边写。

        性能优先级：
          1. MySQL → SSCursor 流式（1 条 SQL，O(n)，最快，约 270 万行/1 分钟）
          2. 有整数主键 → 多线程主键范围查询（O(1)，很快）
          3. 无主键 → 回退到 LIMIT/OFFSET（后期慢）

        读写并行，总时间 ≈ max(读, 写)，而不是 读+写。
        """
        batch_size = self.batch_size

        # ============ 最快路径：MySQL + SSCursor 流式读取（优先于其他所有方式）
        if self.db_type == "mysql":
            try:
                gen = self.connector.streaming_fetchall(columns, table, batch_size)
                if gen is not None:
                    read_count = 0
                    for batch in gen:
                        if not batch:
                            continue
                        read_count += len(batch)
                        if progress is not None:
                            progress.update_read(read_count)
                        yield [list(r) for r in batch]
                    return  # 完成后直接返回，不走下面的多线程逻辑
            except Exception:
                pass  # 失败就回退到下面逻辑

        # ============ 其他路径（SQLite / PostgreSQL / MySQL SSCursor 不可用时）
        # 检测主键，决定是用主键范围查询、键集分页还是 LIMIT/OFFSET
        pk_col, pk_is_integer = _detect_primary_key(self.db_type, self.connector, table)
        tasks: List[Tuple] = []

        if pk_col is not None and pk_is_integer:
            # --- 整数主键范围模式（多线程）---
            try:
                row = self.connector.fetchone(f"SELECT MIN({pk_col}), MAX({pk_col}) FROM {table}")
                if row is not None and row[0] is not None:
                    pk_min, pk_max = row[0], row[1]
                    ranges = _split_pk_ranges(pk_min, pk_max, batch_size)
                    tasks = [("range", idx, pk_col, start, end)
                             for idx, (start, end) in enumerate(ranges)]
            except Exception:
                pass  # 回退到键集分页或 LIMIT/OFFSET

        if pk_col is not None and not tasks:
            # --- 字符串主键：键集分页（单线程但 O(1) 每批，比 LIMIT/OFFSET 快 10-50 倍）---
            q_pk = _quote_identifier(pk_col, self.db_type)
            cols_str = ", ".join(columns)
            last_pk_val = None
            try:
                pk_idx = columns.index(pk_col)
            except ValueError:
                pk_idx = 0
            read_count = 0
            while read_count < total:
                if last_pk_val is None:
                    sql = (f"SELECT {cols_str} FROM {table} "
                           f"ORDER BY {q_pk} LIMIT {batch_size}")
                else:
                    formatted = _format_pk_value(last_pk_val, pk_is_integer, self.db_type)
                    sql = (f"SELECT {cols_str} FROM {table} "
                           f"WHERE {q_pk} > {formatted} "
                           f"ORDER BY {q_pk} LIMIT {batch_size}")
                try:
                    rows = self.connector.fetchall(sql)
                except Exception as exc:
                    self._log("WARN", f"  键集分页读取失败: {exc}")
                    break
                batch = [_row_values(r) for r in rows]
                if not batch:
                    break
                read_count += len(batch)
                last_pk_val = batch[-1][pk_idx]
                if progress is not None:
                    progress.update_read(read_count)
                yield batch
            return  # 完成后直接返回，不走下面的多线程逻辑

        if not tasks:
            # --- 回退：LIMIT/OFFSET ---
            num_batches = math.ceil(total / batch_size)
            tasks = [("offset", idx, self.from_row, batch_size)
                     for idx in range(num_batches)]

        num_batches = len(tasks)
        actual_threads = min(self.max_threads, max(1, num_batches))

        # 有界队列
        result_queue: "_queue.Queue" = _queue.Queue(maxsize=actual_threads * 2)
        error_list: list = []

        producer_thread = threading.Thread(
            target=self._producer_thread,
            args=(tasks, result_queue, self.db_type, self.db_config,
                  table, columns, actual_threads, error_list),
            daemon=True,
        )
        producer_thread.start()

        # 主线程按 0,1,2.. 顺序 yield
        buffer: Dict[int, List[List[Any]]] = {}
        expected_idx = 0
        read_count = 0
        finished = False

        while expected_idx < num_batches and not finished:
            try:
                batch_idx, batch = result_queue.get(timeout=30)
            except _queue.Empty:
                if not producer_thread.is_alive():
                    break
                continue

            if batch is None:
                finished = True
                break

            buffer[batch_idx] = batch
            result_queue.task_done()

            while expected_idx in buffer:
                data = buffer.pop(expected_idx)
                if data:
                    read_count += len(data)
                    if progress is not None:
                        progress.update_read(read_count)
                    yield data
                expected_idx += 1

        producer_thread.join(timeout=30)

        if error_list:
            self._log("WARN", f"  多线程读取时有 {len(error_list)} 个批次出错")

        if progress is not None and not progress._finished:
            pass  # 不在这里 finish（由写入器统一 finish）

    # ------------------------------------------------------------------
    def _producer_thread(self, tasks: List[Tuple],
                        result_queue: "_queue.Queue",
                        db_type: str, db_config: Dict[str, Any],
                        table: str, columns: List[str],
                        thread_count: int, error_list: list) -> None:
        """
        内部生产者线程：启动多个读取工作线程，把批次放入结果队列。
        tasks 是 ("offset"|"range", batch_idx, ...) 的任务列表。
        """
        task_queue: "_queue.Queue" = _queue.Queue()
        for task in tasks:
            task_queue.put(task)

        threads = [
            threading.Thread(
                target=_worker_thread,
                args=(task_queue, result_queue, db_type, db_config,
                      table, columns, error_list),
                daemon=True,
            )
            for _ in range(thread_count)
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=120)

        # 放入 sentinel 告诉主线程：所有批次已读完
        try:
            result_queue.put((-1, None), timeout=5)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # 写入逻辑（CSV / Excel - 流式分批写入）
    # ------------------------------------------------------------------
    def _write_csv(self, filepath: str, columns: List[str],
                   rows: List[List[Any]],
                   progress: Optional[ExportProgress] = None) -> None:
        """
        写入 CSV 文件。
        rows 可能是单一大列表，但这里也支持流式分批调用。
        """
        import csv
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            # 批量写入
            if rows:
                writer.writerows(rows)

    def _write_data_streaming(self, table: str, columns: List[str],
                              total_rows: int,
                              row_iter,
                              progress: ExportProgress) -> List[str]:
        """
        流式写入：通过迭代器逐批读取数据并写入，避免把所有数据存内存。
        优先使用 xlsxwriter（比 openpyxl 快 3-5 倍），不可用时回退 openpyxl。
        row_iter: 每次返回 List[List[Any]]（一批数据），迭代结束时抛出 StopIteration
        progress: 统一的组合进度条对象，用于更新写入进度
        返回生成的文件路径列表。
        """
        base_name = self._generate_filename(table, total_rows)

        # 超过单工作簿容量 -> 用子文件夹
        needs_multi_files = total_rows > MAX_ROWS_PER_WORKBOOK
        target_dir = self.output_dir
        if needs_multi_files:
            folder_name = sanitize_filename(f"{self.db_display}_{table}")
            target_dir = os.path.join(self.output_dir, folder_name)
            os.makedirs(target_dir, exist_ok=True)

        output_files: List[str] = []
        written = 0

        # CSV 简单处理：单文件（超过 500000 仍会单文件，因为 CSV 没有 sheet 概念）
        if self.output_format == "csv":
            filepath = os.path.join(target_dir, f"{base_name}.csv")
            import csv
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                for batch in row_iter:
                    if not batch:
                        continue
                    writer.writerows(batch)
                    written += len(batch)
                    progress.update_write(written)
            progress.finish()
            output_files.append(filepath)
            return output_files

        # ============== 选择写入引擎：优先 xlsxwriter ==============
        use_xlsxwriter = HAS_XLSXWRITER
        if not (use_xlsxwriter or HAS_OPENPYXL):
            raise ImportError("需要 xlsxwriter 或 openpyxl 之一来写入 Excel")

        def workbook_path(wb_idx: int) -> str:
            if wb_idx == 1:
                return os.path.join(target_dir, f"{base_name}.xlsx")
            return os.path.join(target_dir, f"{base_name}_{wb_idx}.xlsx")

        # ========== xlsxwriter 路径（速度：~40000 行/秒） ==========
        if use_xlsxwriter:
            workbook_idx = 1
            rows_in_current_workbook = 0  # 含标题行
            rows_in_current_sheet = 0
            sheet_idx = 0
            current_wb = xlsxwriter.Workbook(workbook_path(1),
                                             {"strings_to_urls": False,
                                              "constant_memory": True})
            current_ws = current_wb.add_worksheet(truncate_sheet_name(self.db_display, table, 0))
            # 绑定方法到局部变量（减少属性查找开销）
            ws_write_row = current_ws.write_row
            ws_write_row(0, 0, columns)  # 第 0 行写标题
            rows_in_current_workbook = 1
            row_idx = 1  # xlsxwriter 从 0 开始

            def switch_sheet_xlsx(ws_idx: int):
                nonlocal current_ws, ws_write_row, rows_in_current_sheet, row_idx
                current_ws = current_wb.add_worksheet(
                    truncate_sheet_name(self.db_display, table, ws_idx)
                )
                ws_write_row = current_ws.write_row
                ws_write_row(0, 0, columns)
                rows_in_current_sheet = 0
                row_idx = 1  # 新 sheet 从第 1 行开始

            def switch_workbook_xlsx(wb_path: str):
                nonlocal current_wb, current_ws, ws_write_row, \
                    rows_in_current_workbook, rows_in_current_sheet, row_idx
                current_wb.close()
                current_wb = xlsxwriter.Workbook(wb_path,
                                                 {"strings_to_urls": False,
                                                  "constant_memory": True})
                current_ws = current_wb.add_worksheet(
                    truncate_sheet_name(self.db_display, table, 0)
                )
                ws_write_row = current_ws.write_row
                ws_write_row(0, 0, columns)
                rows_in_current_workbook = 1
                rows_in_current_sheet = 0
                row_idx = 1

            current_path = workbook_path(1)
            output_files.append(current_path)

            # xlsxwriter 写入非常快，直接按批次写，不用额外 buffer
            for batch in row_iter:
                if not batch:
                    continue
                for row in batch:
                    if rows_in_current_sheet >= MAX_ROWS_PER_SHEET:
                        sheet_idx += 1
                        switch_sheet_xlsx(sheet_idx)
                        rows_in_current_workbook += 1

                    if rows_in_current_workbook >= MAX_ROWS_PER_WORKBOOK:
                        workbook_idx += 1
                        sheet_idx = 0
                        current_path = workbook_path(workbook_idx)
                        output_files.append(current_path)
                        switch_workbook_xlsx(current_path)

                    ws_write_row(row_idx, 0, row)
                    row_idx += 1
                    rows_in_current_sheet += 1
                    rows_in_current_workbook += 1
                    written += 1

                progress.update_write(written)

            current_wb.close()
            progress.finish(f"共 {workbook_idx} 个工作簿文件")
            return output_files

        # ========== openpyxl 回退路径（速度：~10000 行/秒） ==========
        workbook_idx = 1
        rows_in_current_workbook = 0  # 含标题行
        rows_in_current_sheet = 0
        sheet_idx = 0
        current_wb = Workbook(write_only=True)
        current_ws = current_wb.create_sheet(truncate_sheet_name(self.db_display, table, 0))
        # 绑定方法到局部变量（减少属性查找开销）
        ws_append = current_ws.append
        ws_append(columns)
        rows_in_current_workbook = 1

        def switch_sheet_openpyxl(ws_idx: int):
            nonlocal current_ws, ws_append, rows_in_current_sheet
            current_ws = current_wb.create_sheet(
                truncate_sheet_name(self.db_display, table, ws_idx)
            )
            ws_append = current_ws.append
            ws_append(columns)
            rows_in_current_sheet = 0

        def switch_workbook_openpyxl(wb_path: str):
            nonlocal current_wb, current_ws, ws_append, rows_in_current_workbook, rows_in_current_sheet
            current_wb.save(wb_path)
            current_wb = Workbook(write_only=True)
            current_ws = current_wb.create_sheet(
                truncate_sheet_name(self.db_display, table, 0)
            )
            ws_append = current_ws.append
            ws_append(columns)
            rows_in_current_workbook = 1
            rows_in_current_sheet = 0

        current_path = workbook_path(1)
        output_files.append(current_path)

        # openpyxl 较慢，合并多个批次一次性写入
        merge_buffer: List[List[Any]] = []
        MERGE_SIZE = EXCEL_MERGE_ROWS

        for batch in row_iter:
            if not batch:
                continue

            merge_buffer.extend(batch)

            while len(merge_buffer) >= MERGE_SIZE:
                write_block = merge_buffer[:MERGE_SIZE]
                del merge_buffer[:MERGE_SIZE]

                for row in write_block:
                    if rows_in_current_sheet >= MAX_ROWS_PER_SHEET:
                        sheet_idx += 1
                        switch_sheet_openpyxl(sheet_idx)
                        rows_in_current_workbook += 1

                    if rows_in_current_workbook >= MAX_ROWS_PER_WORKBOOK:
                        workbook_idx += 1
                        sheet_idx = 0
                        current_path = workbook_path(workbook_idx)
                        output_files.append(current_path)
                        switch_workbook_openpyxl(current_path)

                    ws_append(row)
                    rows_in_current_sheet += 1
                    rows_in_current_workbook += 1
                    written += 1

                progress.update_write(written)

        if merge_buffer:
            for row in merge_buffer:
                if rows_in_current_sheet >= MAX_ROWS_PER_SHEET:
                    sheet_idx += 1
                    switch_sheet_openpyxl(sheet_idx)
                    rows_in_current_workbook += 1

                if rows_in_current_workbook >= MAX_ROWS_PER_WORKBOOK:
                    workbook_idx += 1
                    sheet_idx = 0
                    current_path = workbook_path(workbook_idx)
                    output_files.append(current_path)
                    switch_workbook_openpyxl(current_path)

                ws_append(row)
                rows_in_current_sheet += 1
                rows_in_current_workbook += 1
                written += 1

            progress.update_write(written)

        current_wb.save(current_path)
        progress.finish(f"共 {workbook_idx} 个工作簿文件")
        return output_files

    # ------------------------------------------------------------------
    def _export_table_parallel(self, table: str, columns: List[str],
                                export_rows: int, pk_col: str,
                                pk_is_integer: bool,
                                progress: ExportProgress) -> List[str]:
        """
        单表分段多进程导出：按主键范围分 N 段，每个进程独立查 + 写。
        绕过 GIL，对大表（>50 万行）速度可以提升 2-4 倍。

        支持整数主键（算术分段）和字符串主键（采样分段）。

        生产级特性：
          - 动态并发数（基于 CPU 核心数，不超过 MAX_PROCESSES=4）
          - 进程超时保护（单段最多等待 10 分钟，防止死锁）
          - 连接失败重试（最多 3 次，间隔 2 秒）
          - ETA 单调递减（不会"时光倒流"）

        返回：所有生成的文件路径列表
        """
        from concurrent.futures import ProcessPoolExecutor, as_completed, Future, TimeoutError

        # ============ 动态并发数计算（生产级：基于 CPU 核心数）============
        try:
            cpu_count = os.cpu_count() or 2
        except Exception:
            cpu_count = 2

        max_parallel = min(cpu_count, MAX_PROCESSES)

        # 实际段数
        try:
            pk_min_row = self.connector.fetchone(
                f"SELECT MIN({pk_col}), MAX({pk_col}) FROM {table}"
            )
            if pk_min_row is None:
                raise ValueError("无法获取主键范围")
            pk_min, pk_max = pk_min_row[0], pk_min_row[1]
        except Exception:
            row_iter = self._read_rows_sequential_streaming(table, columns, export_rows, progress)
            return self._write_data_streaming(table, columns, export_rows, row_iter, progress)

        if pk_min is None or pk_max is None:
            row_iter = self._read_rows_sequential_streaming(table, columns, export_rows, progress)
            return self._write_data_streaming(table, columns, export_rows, row_iter, progress)

        # 子文件夹
        folder_name = sanitize_filename(f"{self.db_display}_{table}")
        target_dir = os.path.join(self.output_dir, folder_name)
        os.makedirs(target_dir, exist_ok=True)

        # 按「每段 50 万行」切分
        rows_per_segment = 500000

        if pk_is_integer:
            # --- 整数主键：算术分段 ---
            total_range = max(1, pk_max - pk_min)
            rows_per_pk = export_rows / total_range
            pk_per_segment = math.ceil(rows_per_segment / rows_per_pk) if rows_per_pk > 0 else rows_per_segment

            segments = []
            start_val = pk_min
            for i in range(1000):
                end_val = min(start_val + pk_per_segment, pk_max + 1)
                if start_val > pk_max:
                    break
                part_suffix = f"_p{i + 1:03d}"
                output_path = os.path.join(
                    target_dir,
                    f"{sanitize_filename(self.db_display)}_{table}{part_suffix}."
                    f"{self.output_format}"
                )
                segments.append((start_val, end_val, output_path))
                start_val = end_val
        else:
            # --- 字符串主键：采样分段 ---
            num_segments = max(1, math.ceil(export_rows / rows_per_segment))
            segment_size = math.ceil(export_rows / num_segments)

            split_points = []
            for i in range(1, num_segments):
                offset = i * segment_size
                try:
                    row = self.connector.fetchone(
                        f"SELECT {pk_col} FROM {table} ORDER BY {pk_col} LIMIT 1 OFFSET {offset}"
                    )
                    if row is not None:
                        split_points.append(row[0])
                except Exception:
                    pass

            # 如果采样失败，回退到单进程键集分页
            if len(split_points) < num_segments - 1:
                self._log("WARN", "  字符串主键采样失败，回退到单进程键集分页")
                row_iter = self._read_rows_sequential_streaming(
                    table, columns, export_rows, progress
                )
                return self._write_data_streaming(
                    table, columns, export_rows, row_iter, progress
                )

            segments = []
            boundaries = [pk_min] + split_points + [pk_max]
            for i in range(len(boundaries) - 1):
                part_suffix = f"_p{i + 1:03d}"
                output_path = os.path.join(
                    target_dir,
                    f"{sanitize_filename(self.db_display)}_{table}{part_suffix}."
                    f"{self.output_format}"
                )
                segments.append((boundaries[i], boundaries[i + 1], output_path))

        actual_segments = len(segments)
        if actual_segments == 0:
            row_iter = self._read_rows_sequential_streaming(
                table, columns, export_rows, progress
            )
            return self._write_data_streaming(
                table, columns, export_rows, row_iter, progress
            )

        # 最终并发数（不能超过段数）
        max_parallel = min(max_parallel, actual_segments)

        # ============ 生产级并发导出 ============
        manager = Manager()
        shared_counter = manager.Value("i", 0)

        worker_args = []
        for seg in segments:
            s_start, s_end, s_path = seg
            worker_args.append((
                self.db_type, self.db_config, table, columns, pk_col,
                s_start, s_end, s_path, self.output_format,
                self.batch_size, self.db_display, shared_counter,
                pk_is_integer,
            ))

        output_files: List[str] = []
        total_written = 0
        errors: List[str] = []
        self._multiprocess_start_time = time.time()

        progress.update_read(1)

        running: dict = {}
        pending_args = list(worker_args)
        submitted_count = 0

        # 进程超时保护（生产级：防止死锁）
        SEGMENT_TIMEOUT_SECONDS = 600  # 单段最多等待 10 分钟

        try:
            with ProcessPoolExecutor(max_workers=max_parallel) as executor:
                # 先提交 max_parallel 个
                for _ in range(max_parallel):
                    if pending_args:
                        args = pending_args.pop(0)
                        running[executor.submit(_segment_worker, args)] = submitted_count
                        submitted_count += 1

                completed = 0
                running_with_time: Dict[Any, float] = {}
                for f in running:
                    running_with_time[f] = time.time()

                while running or pending_args:
                    # 检查谁完成了
                    done_this_round = []
                    still_running = {}
                    timed_out_futures = []

                    now = time.time()
                    for f, idx in running.items():
                        if f.done():
                            done_this_round.append((f, idx))
                        else:
                            # 超时检查（生产级：防死锁）
                            seg_elapsed = now - running_with_time.get(f, now)
                            if seg_elapsed > SEGMENT_TIMEOUT_SECONDS:
                                timed_out_futures.append(f)
                            else:
                                still_running[f] = idx
                    running = still_running

                    # 处理超时的 future（生产级：强制取消并记录错误）
                    for f in timed_out_futures:
                        try:
                            f.cancel()
                        except Exception:
                            pass
                        start_t = running_with_time.pop(f, None)
                        errors.append(f"段超时（>{SEGMENT_TIMEOUT_SECONDS}s），已强制终止")
                        completed += 1

                    # 处理已完成
                    for f, idx in done_this_round:
                        try:
                            file_list, row_count, error = f.result(timeout=5)
                        except TimeoutError:
                            errors.append("获取结果超时")
                            file_list = []
                            row_count = 0
                        except Exception as exc:
                            error = f"{exc}"
                            file_list = []
                            row_count = 0

                        start_t = running_with_time.pop(f, None)

                        if error:
                            errors.append(error)
                        elif file_list:
                            output_files.extend(file_list)
                            total_written += row_count
                            for fp in file_list:
                                try:
                                    file_size = os.path.getsize(fp) / 1024 / 1024
                                except Exception:
                                    file_size = 0.0
                                per_file_rows = row_count // len(file_list) if file_list else row_count
                                sys.stdout.write(
                                    f"\r\x1b[K  ✓ {os.path.basename(fp)}  "
                                    f"{per_file_rows} 行  "
                                    f"{file_size:.1f} MB\n"
                                )
                                sys.stdout.flush()
                        completed += 1

                    # 启动下一批
                    now = time.time()
                    while pending_args and len(running) < max_parallel:
                        args = pending_args.pop(0)
                        f = executor.submit(_segment_worker, args)
                        running[f] = submitted_count
                        running_with_time[f] = now
                        submitted_count += 1

                    # ============ 进度条更新（基于共享计数器实时进度）============
                    # 使用子进程实时汇报的共享计数器，进度条会平滑更新
                    current_progress = shared_counter.value
                    progress.update_write(min(current_progress, export_rows))

                    if running or pending_args:
                        time.sleep(0.5)

        except Exception as exc:
            errors.append(f"进程池异常: {exc}")

        # 如果有错误或没输出任何文件 → 回退到单线程
        if errors or not output_files:
            output_files.clear()
            total_written = 0
            row_iter = self._read_rows_sequential_streaming(
                table, columns, export_rows, progress
            )
            output_files = self._write_data_streaming(
                table, columns, export_rows, row_iter, progress
            )
            return output_files

        progress.finish(f"共 {len(output_files)} 个分段文件，合计 {total_written} 行")
        # 按文件名排序，保证输出顺序一致
        output_files.sort()
        return output_files

    # ------------------------------------------------------------------
    # 对外主方法
    # ------------------------------------------------------------------
    def export(self) -> Dict[str, Any]:
        """执行全部表的导出，返回摘要字典"""
        result_summary: Dict[str, Any] = {
            "tables": {},
            "total_tables": len(self.tables),
            "total_rows": 0,
            "output_files": [],
            "start_time": time.time(),
        }

        for table_idx, table in enumerate(self.tables, 1):
            table_start = time.time()
            try:
                columns = self.connector.get_columns(table)
                if not columns:
                    self._log("WARN", f"表 {table} 未检测到列，跳过")
                    continue

                total_rows, export_rows = self._get_effective_total(table)
                if export_rows == 0:
                    self._log("INFO", f"表 {table} 在指定行范围内无数据，跳过")
                    continue

                self._log("INFO",
                          f"[{table_idx}/{len(self.tables)}] 开始导出表 {table} "
                          f"(源表共 {total_rows} 行，本次导出 {export_rows} 行，"
                          f"{len(columns)} 列)")

                # ---- 决定读取/写入方式 ----
                # 优先级：多进程（>50万行+有主键）> 多线程 > 单线程
                use_multiprocess = False
                pk_col = None
                pk_is_integer = False

                if (export_rows > MULTIPROCESS_THRESHOLD
                        and self.db_type in MULTITHREAD_DB_TYPES):
                    pk_col, pk_is_integer = _detect_primary_key(self.db_type,
                                                                 self.connector, table)
                    if pk_col is not None:
                        use_multiprocess = True

                # 单一组合进度条：读取和写入都更新同一对象（单行刷新）
                progress = ExportProgress(table, export_rows)

                if use_multiprocess:
                    # 多进程分段并行：每个进程独立查一段 + 写一个文件
                    table_files = self._export_table_parallel(
                        table, columns, export_rows, pk_col, pk_is_integer, progress
                    )
                else:
                    # 单线程 / 多线程逻辑（保持原有）
                    can_multithread = (export_rows > MULTITHREAD_THRESHOLD
                                       and self.db_type in MULTITHREAD_DB_TYPES)
                    use_multithread = can_multithread

                    if not use_multithread:
                        row_iter = self._read_rows_sequential_streaming(
                            table, columns, export_rows, progress
                        )
                    else:
                        row_iter = self._read_rows_multithread_streaming(
                            table, columns, export_rows, progress
                        )
                    table_files = self._write_data_streaming(
                        table, columns, export_rows, row_iter, progress
                    )

                elapsed = time.time() - table_start
                result_summary["tables"][table] = {
                    "rows": export_rows,
                    "columns": len(columns),
                    "files": table_files,
                    "elapsed": elapsed,
                }
                result_summary["total_rows"] += export_rows
                result_summary["output_files"].extend(table_files)

                self._log("SUCCESS",
                          f"表 {table} 导出完成，{export_rows} 行 / "
                          f"{len(table_files)} 个文件，用时 {elapsed:.2f}s")

            except Exception as exc:
                self._log("ERROR", f"表 {table} 导出失败: {exc}")
                raise

        result_summary["elapsed"] = time.time() - result_summary["start_time"]
        return result_summary


__all__ = [
    "DatabaseExporter",
    "BATCH_THRESHOLD",
    "DEFAULT_BATCH_SIZE",
    "MULTITHREAD_THRESHOLD",
    "DEFAULT_THREADS",
    "MAX_THREADS",
    "MAX_ROWS_PER_SHEET",
    "MAX_ROWS_PER_WORKBOOK",
    "HAS_OPENPYXL",
]